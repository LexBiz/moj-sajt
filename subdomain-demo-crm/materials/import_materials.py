#!/usr/bin/env python3
"""
Material catalog import — reads KV Elektro XLSX + Kundenpreise XLSX
and inserts into crm_service_catalog_items (itemKind=material)
"""
import os, sys, json, re
import openpyxl
import psycopg2
import psycopg2.extras

DATABASE_URL = os.environ.get('DATABASE_URL') or os.environ.get('CRM_DATABASE_URL')
if not DATABASE_URL:
    print('ERROR: DATABASE_URL not set'); sys.exit(1)

MATERIALS_DIR = os.environ.get('MATERIALS_DIR', os.path.dirname(__file__))

def parse_unit(baleni):
    s = str(baleni or '').strip()
    parts = s.split()
    if len(parts) >= 2:
        return parts[-1].lower()
    return 'ks'

def clean_price(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return round(float(v), 4)
    s = re.sub(r'[\s\xa0]', '', str(v)).replace(',', '.')
    try: return round(float(s), 4)
    except: return 0.0

def bulk_insert(conn, items, batch_size=1000):
    if not items: return 0
    cur = conn.cursor()
    inserted = 0
    for i in range(0, len(items), batch_size):
        chunk = items[i:i+batch_size]
        records = [
            (
                it.get('source_code'),
                it['item_name'],
                it.get('item_description'),
                it.get('unit', 'ks'),
                it.get('base_price', 0),
                'electro',           # trade_type
                it.get('category_key', 'elektro'),
                it.get('phase_key', 'material'),
                True,
                json.dumps(it.get('metadata', {})),
            )
            for it in chunk
        ]
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO crm_service_catalog_items
               (source_code, item_name, item_description, unit, base_price,
                trade_type, category_key, phase_key, is_active, metadata,
                created_at, updated_at)
               VALUES %s
               ON CONFLICT (source_code) WHERE source_code IS NOT NULL
               DO UPDATE SET
                 item_name = EXCLUDED.item_name,
                 item_description = EXCLUDED.item_description,
                 unit = EXCLUDED.unit,
                 base_price = EXCLUDED.base_price,
                 metadata = EXCLUDED.metadata,
                 updated_at = now()""",
            records,
            template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now(),now())",
            page_size=batch_size
        )
        conn.commit()
        inserted += len(chunk)
        print(f'\r  inserted {inserted}/{len(items)}...', end='', flush=True)
    print()
    cur.close()
    return inserted

# ── KV Elektro XLSX ──
# Cols: Kód produktu, Produkt, Kategorie, Značka, Balení, EAN, Kód dodavatele, Základní cena, Moje cena
def import_kv_elektro(conn):
    fpath = os.path.join(MATERIALS_DIR, '0002053593_pricelist KV ELektro.xlsx')
    print(f'\n📗 KV Elektro XLSX: {fpath}')
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    items = []
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code    = str(row[0] or '').strip()
        name    = str(row[1] or '').strip()
        cat     = str(row[2] or '').strip()
        brand   = str(row[3] or '').strip()
        baleni  = str(row[4] or '')
        sup_cod = str(row[6] or '').strip()
        moje    = clean_price(row[8]) if len(row) > 8 else 0  # Moje cena
        base    = clean_price(row[7]) if len(row) > 7 else 0
        price   = moje if moje > 0 else base
        if not name: continue
        unit = parse_unit(baleni)
        full_name = f'{name} [{brand}]' if brand else name
        items.append({
            'source_code': code or sup_cod or None,
            'item_name': full_name[:400],
            'item_description': cat[:255] if cat else None,
            'unit': unit,
            'base_price': price,
            'category_key': 'elektro',
            'phase_key': 'material',
            'metadata': {
                'itemKind': 'material',
                'category': 'elektro',
                'brand': brand,
                'supplierCode': sup_cod,
                'catalogCategory': cat,
            }
        })
        count += 1
        if count % 20000 == 0:
            print(f'  read {count} rows...', flush=True)
    wb.close()
    print(f'  total read: {len(items)} items')
    n = bulk_insert(conn, items)
    print(f'✅ KV Elektro: {n} položek importováno')
    return n

# ── Kundenpreise XLSX ──
# Cols: Artikel, RGR, Bezeichung, PEH, Preisbasis, Rabatt, konová cena
def import_kundenpreise(conn):
    fpath = os.path.join(MATERIALS_DIR, '20251014-Kundenpreise-407865.xlsx')
    print(f'\n📗 Kundenpreise XLSX: {fpath}')
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    items = []
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code  = str(row[0] or '').strip()
        name  = str(row[2] or '').strip()
        price = clean_price(row[6])  # konová cena
        if not name or not code: continue
        items.append({
            'source_code': f'KP-{code}',
            'item_name': name[:400],
            'item_description': None,
            'unit': 'ks',
            'base_price': price,
            'category_key': 'elektro',
            'phase_key': 'material',
            'metadata': {
                'itemKind': 'material',
                'category': 'elektro',
                'artikelCode': code,
            }
        })
        count += 1
        if count % 10000 == 0:
            print(f'  read {count} rows...', flush=True)
    wb.close()
    print(f'  total read: {len(items)} items')
    n = bulk_insert(conn, items)
    print(f'✅ Kundenpreise: {n} položek importováno')
    return n

def main():
    print('🔌 Connecting to database...')
    conn = psycopg2.connect(DATABASE_URL)
    print('✅ Connected')
    total = 0
    try:
        total += import_kv_elektro(conn)
    except Exception as e:
        print(f'❌ KV Elektro error: {e}')
    try:
        total += import_kundenpreise(conn)
    except Exception as e:
        print(f'❌ Kundenpreise error: {e}')
    conn.close()
    print(f'\n🎉 TOTAL IMPORTED: {total} materiálů')

if __name__ == '__main__':
    main()
